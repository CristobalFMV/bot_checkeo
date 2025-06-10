
from openpyxl import load_workbook
import checkeo
from ruta_excel import archivo_vicarius

archivo_vicarius = archivo_vicarius
wb = load_workbook(archivo_vicarius)
ws = wb.active
ip_local = checkeo.checkIP()


def escribirExcel(ws, fila):
    ws[f"C{fila}"] = checkeo.checkCurrentUser()
    ws[f"D{fila}"] = checkeo.pedirPassword()
    ws[f"E{fila}"] = checkeo.pedirAnexo()
    ws[f"F{fila}"] = checkeo.checkSerial()
    ws[f"G{fila}"] = checkeo.checkMac()
    ws[f"H{fila}"] = checkeo.pedirUbicacion()
    ws[f"I{fila}"] = checkeo.checkHost()
    ws[f"J{fila}"] = checkeo.check_app_instalada("Topia.exe")
    ws[f"K{fila}"] = checkeo.checkFirewall()
    ws[f"L{fila}"] = checkeo.check_app_instalada("Sophos UI.exe")
    ws[f"M{fila}"] = checkeo.checkOS()
    ws[f"N{fila}"] = checkeo.checkTipoEquipo()
    ws[f"O{fila}"] = checkeo.checkScreen()
    ws[f"P{fila}"] = checkeo.tieneHuellero()
    ws[f"Q{fila}"] = checkeo.checkYTB()
    ws[f"R{fila}"] = checkeo.checkAdmin()
    ws[f"S{fila}"] = checkeo.checkRDP()
    ws[f"T{fila}"] = checkeo.checkObservaciones()
    ws[f"U{fila}"] = checkeo.checkNombreSoporte()
    ws[f"V{fila}"] = checkeo.checkFecha()

