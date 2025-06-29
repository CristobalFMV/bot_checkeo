from openpyxl import load_workbook
from core.modules import checkeo
from core.config.ruta_excel import archivo_vicarius
archivo_vicarius = archivo_vicarius
wb = load_workbook(archivo_vicarius)
ws = wb.active
ip_local = checkeo.checkIP()
'''Parametros para el SQL'''
'''users,password,anexo,
                serial,mac,ubicacion,host,
                app_install1,firewall,app_install2,
                sistema_op, tipo_cpu,pantalla,huellero,
                ytb_premium, admin,remoto,
                observacion,nombre_soporte,fecha'''
#============================================
"""Objeto equipo que almacenara la información por equipo tomado"""
class Equipo:
    def __init__(
        self,
        ip,
        usuario,
        anexo,
        serial,
        mac,
        ubicacion,
        host,
        app_install1,
        firewall,
        app_install2,
        sistema_op,
        tipo_cpu,
        pantalla,
        huellero,
        ytb_premium,
        admin,
        remoto,
        observacion,
        nombre_soporte,
        fecha,
        id=None,
    ):
        self.id = id
        self.ip = ip
        self.users = usuario
        self.anexo = anexo
        self.serial = serial
        self.mac = mac
        self.ubicacion = ubicacion
        self.host = host
        self.app_install1 = app_install1
        self.firewall = firewall
        self.app_install2 = app_install2
        self.sistema_op = sistema_op
        self.tipo_cpu = tipo_cpu
        self.pantalla = pantalla
        self.huellero = huellero
        self.ytb_premium = ytb_premium
        self.admin = admin
        self.remoto = remoto
        self.observacion = observacion
        self.nombre_soporte = nombre_soporte
        self.fecha = fecha

    def escribir_en_excel(self, ws, fila):
        columnas = ["C", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V"]
        data_equipo = [
            self.users, self.anexo, self.serial, self.mac, self.ubicacion, self.host,
            self.app_install1, self.firewall, self.app_install2, self.sistema_op,
            self.tipo_cpu, self.pantalla, self.huellero, self.ytb_premium, self.admin,
            self.remoto, self.observacion, self.nombre_soporte, self.fecha
        ]
        try:
            for letra, data in zip(columnas, data_equipo):
                ws[f"{letra}{fila}"] = data
        except Exception as e:
            print(f"Error escribiendo en Excel: {e}")
            print("❌ Error durante la escritura en Excel:", e)
        else:
            print("✅ Proceso completado con éxito.")




#equipo = Equipo()
#datos = equipo.__dict__